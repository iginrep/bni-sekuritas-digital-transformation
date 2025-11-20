# scrap_to_excel_technical.py
# Requirements: requests, pandas, openpyxl, matplotlib, wordcloud, nltk, pillow

import sys
import subprocess
import importlib

# --- Install otomatis paket yang diperlukan ---
required = [
    "requests", "pandas", "openpyxl", "matplotlib",
    "wordcloud", "nltk", "Pillow"
]

def install(pkg):
    subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])

for pkg in required:
    try:
        importlib.import_module(pkg)
    except Exception:
        print(f"Installing {pkg}...")
        install(pkg)

# --- Import setelah install ---
import requests
import pandas as pd
import matplotlib.pyplot as plt
from wordcloud import WordCloud
import nltk
from nltk.sentiment.vader import SentimentIntensityAnalyzer
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import os
import re

nltk.download("vader_lexicon", quiet=True)

# --- Konfigurasi ---
APP_ID = "1469177051"      # BIONS Mobile
COUNTRY = "id"
OUT_XLSX = "bions_reviews_full.xlsx"
WORDCLOUD_PNG = "bions_wordcloud_technical.png"
RATING_PNG = "bions_rating_dist.png"

# --- Stopwords Bahasa Indonesia + Custom ---
stopwords_basic = {
    "dan","yang","di","ke","dari","dengan","untuk","pada","itu","ini","saya","kita",
    "kami","juga","atau","karena","jadi","agar","dll","tp","gmna","aja","lagi",
    "sih","nya","ada","tidak","gak","ga","udah","baik","banget","bgt","kayak",
    "mau","tolong","tolonglah","tolongin","buat","semua","setiap","ketika", "saat", "tapi",
    "lebih","kurang","bisa","bisaaa","gitu","nih", "maluin", "cuih", "lagin", "duit", "apalagi"
}

# Stopwords tambahan non-teknis
stopwords_custom = {
    "aplikasi","aplikasinya","bions","bni","sekuritas","saham","user","investor",
    "order","beli","jual","jual","menu","tampilan","akun","fitur","versi","update",
    "bagus","kurang","telat","proses","harga", "market","chart","watchlist"
}

STOPWORDS = stopwords_basic.union(stopwords_custom)

# --- Kata teknis yang ingin difokuskan ---
technical_keywords = [
    "error","bug","login","gagal","crash","close","force","lemot","lambat",
    "loading","blank","hang","server","network","otp","data","system","sistem",
    "delay","timeout","disconnect","lag","password","otp","api","ui","ux",
    "security","risiko","update","versi","transaksi","problem","issue","failed", "connect","sync","sinkron","notifikasi","notification","backup","restore", "install","uninstall", "valid", 
]

# Precompile regex untuk mencari kata teknis
technical_pattern = re.compile(r"|".join(technical_keywords), re.IGNORECASE)

# --- Fungsi ambil satu halaman ---
def fetch_page(page):
    url = f"https://itunes.apple.com/{COUNTRY}/rss/customerreviews/page={page}/id={APP_ID}/sortby=mostrecent/json"
    try:
        r = requests.get(url, timeout=10)
        r.raise_for_status()
        return r.json()
    except:
        return None

# --- Scrape multi-page ---
all_reviews = []
page = 1

print("Mulai scraping...")
while True:
    print(f"  Page {page}")
    data = fetch_page(page)

    if not data or "feed" not in data or "entry" not in data["feed"]:
        break

    entries = data["feed"]["entry"]
    if len(entries) <= 1:
        break

    for e in entries[1:]:
        nama = e.get("author", {}).get("name", {}).get("label", "")
        komentar = e.get("content", {}).get("label", "")
        rating = e.get("im:rating", {}).get("label", "")
        judul = e.get("title", {}).get("label", "")
        tanggal = e.get("updated", {}).get("label", "")

        try: r_int = int(rating)
        except: r_int = None

        all_reviews.append({
            "nama": nama,
            "komentar": komentar,
            "bintang": r_int,
            "judul": judul,
            "tanggal": tanggal
        })

    page += 1

# --- DataFrame ---
df = pd.DataFrame(all_reviews)
print("Total review:", len(df))

# --- Sentiment Analysis ---
sia = SentimentIntensityAnalyzer()

def classify_sentiment(text):
    if not isinstance(text, str) or text.strip()=="":
        return "neutral", 0
    vs = sia.polarity_scores(text)
    comp = vs["compound"]
    if comp >= 0.05:  return "positive", comp
    if comp <= -0.05: return "negative", comp
    return "neutral", comp

sent = [classify_sentiment(t) for t in df["komentar"].astype(str)]
df["sentiment"] = [s[0] for s in sent]
df["sentiment_score"] = [s[1] for s in sent]

# --- Simpan Excel ---
with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Reviews", index=False)

    # summary rating
    rc = df["bintang"].value_counts().sort_index()
    rc.rename_axis("bintang").reset_index(name="jumlah").to_excel(
        writer, sheet_name="RatingSummary", index=False
    )

    # summary sentiment
    df["sentiment"].value_counts().rename_axis("sentiment").reset_index(name="jumlah").to_excel(
        writer, sheet_name="SentimentSummary", index=False
    )

# --- Grafik Distribusi Rating ---
plt.figure(figsize=(8,5))
bins = [1,2,3,4,5]
counts = [int(rc.get(i,0)) for i in bins]
plt.bar([str(i) for i in bins], counts)
plt.title("Distribusi Rating BIONS Mobile (App Store)")
plt.xlabel("Bintang")
plt.ylabel("Jumlah Ulasan")
plt.tight_layout()
plt.savefig(RATING_PNG, dpi=200)
plt.close()

# --- WordCloud teknis ---
clean_words = []

for text in df["komentar"].astype(str):
    words = re.findall(r"\b[a-zA-Z0-9]+\b", text.lower())
    for w in words:
        if w in STOPWORDS:
            continue
        if len(w) <= 2:
            continue
        if technical_pattern.search(w):   # hanya kata teknis
            clean_words.append(w)

filtered_text = " ".join(clean_words)

wc = WordCloud(
    width=1600,
    height=800,
    background_color="white",
    colormap="tab20",
    collocations=False,
    max_words=200
).generate(filtered_text)

wc.to_file(WORDCLOUD_PNG)
print("WordCloud teknis tersimpan:", WORDCLOUD_PNG)

# --- Masukkan gambar ke Excel ---
try:
    wb = load_workbook(OUT_XLSX)
    ws = wb.create_sheet("Visuals")

    if os.path.exists(RATING_PNG):
        img1 = XLImage(RATING_PNG)
        img1.anchor = "A1"
        ws.add_image(img1)

    if os.path.exists(WORDCLOUD_PNG):
        img2 = XLImage(WORDCLOUD_PNG)
        img2.anchor = "A25"
        ws.add_image(img2)

    wb.save(OUT_XLSX)
except Exception as e:
    print("Gagal memasukkan gambar ke Excel:", e)

print("\n=== Selesai ===")
print("Excel:", OUT_XLSX)
print("WordCloud:", WORDCLOUD_PNG)
print("Rating Chart:", RATING_PNG)
